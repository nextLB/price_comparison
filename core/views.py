from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from datetime import datetime, date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import (
    User, Department, Role, Permission,
    Customer, Contact, Project, FollowUp,
    Trip, Expense, Drug, Pharmacy, PharmaceuticalCompany,
    PharmacyRecord, CompanyRecord, DrugRecordStatus, AnchorPrice,
    PharmacyUser, CompanyUser, District, PharmacyRecordReview, DrugPriceReview
)
from .price_calculator import extract_number, process_all_drugs


def _safe_str(val):
    """安全转换值为字符串，处理NaN"""
    if val is None:
        return ''
    if isinstance(val, float) and (val != val):
        return ''
    return str(val)


# ==================== 认证相关 ====================

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, '用户名或密码错误')
    return render(request, 'core/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


def register_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        email = request.POST.get('email', '')
        phone = request.POST.get('phone', '')
        user_type = request.POST.get('user_type', 'pharmacy')
        organization = request.POST.get('organization', '')
        medical_insurance_code = request.POST.get('medical_insurance_code', '')

        if User.objects.filter(username=username).exists():
            messages.error(request, '用户名已存在')
            return redirect('register')

        if password != confirm_password:
            messages.error(request, '两次密码输入不一致')
            return redirect('register')

        user = User.objects.create_user(
            username=username, email=email, password=password,
            phone=phone, user_type=user_type, organization=organization,
            medical_insurance_code=medical_insurance_code,
        )

        # 如果是药店或药企用户注册，尝试自动关联
        if user_type == 'pharmacy':
            # 尝试按机构名称查找药店并自动关联
            if organization:
                pharmacy = Pharmacy.objects.filter(pharmacy_name=organization).first()
                if pharmacy:
                    PharmacyUser.objects.get_or_create(user=user, defaults={'pharmacy': pharmacy})
                    messages.info(request, f'已自动关联至药店：{pharmacy.pharmacy_name}')
                else:
                    messages.warning(request, '注册成功！但未找到匹配的药店，请联系药事所监督科进行药店编码关系维护')
        elif user_type == 'company':
            if organization:
                company = PharmaceuticalCompany.objects.filter(company_name=organization).first()
                if company:
                    CompanyUser.objects.get_or_create(user=user, defaults={'company': company})
                    messages.info(request, f'已自动关联至药企：{company.company_name}')
                else:
                    messages.warning(request, '注册成功！但未找到匹配的药企，请联系药事所药品科进行药企编码关系维护')

        messages.success(request, '注册成功，请登录')
        return redirect('login')

    return render(request, 'core/register.html')


# ==================== 首页和仪表盘 ====================

@login_required
def dashboard(request):
    user = request.user
    context = {
        'user': user,
        'drug_count': Drug.objects.count(),
        'pharmacy_count': Pharmacy.objects.count(),
        'company_count': PharmaceuticalCompany.objects.count(),
        'user_count': User.objects.count(),
        'anchor_count': AnchorPrice.objects.count(),
        'pharmacy_record_count': PharmacyRecord.objects.count(),
        'pending_review_count': DrugPriceReview.objects.filter(status=DrugRecordStatus.PENDING).count(),
        'approved_review_count': DrugPriceReview.objects.filter(status=DrugRecordStatus.APPROVED).count(),
        'rejected_review_count': DrugPriceReview.objects.filter(status=DrugRecordStatus.REJECTED).count(),
        'pending_pharmacy_review_count': PharmacyRecord.objects.filter(status=DrugRecordStatus.SUBMITTED).count(),
    }

    # 药店用户专用统计
    if user.user_type == 'pharmacy':
        try:
            pharmacy = user.pharmacy_profile.pharmacy
            context['my_record_count'] = PharmacyRecord.objects.filter(pharmacy=pharmacy).count()
            context['my_approved_count'] = PharmacyRecord.objects.filter(
                pharmacy=pharmacy, status=DrugRecordStatus.APPROVED
            ).count()
        except:
            context['my_record_count'] = 0
            context['my_approved_count'] = 0

    # 药企用户专用统计
    if user.user_type == 'company':
        try:
            company = user.company_profile.company
            context['my_record_count'] = CompanyRecord.objects.filter(company=company).count()
            context['my_approved_count'] = CompanyRecord.objects.filter(
                company=company, status=DrugRecordStatus.APPROVED
            ).count()
        except:
            context['my_record_count'] = 0
            context['my_approved_count'] = 0

    # 区县医保局专用统计
    if user.user_type == 'district':
        district_name = user.organization
        if district_name:
            base_qs = PharmacyRecord.objects.filter(pharmacy__district__icontains=district_name)
            context['pending_review_count'] = base_qs.filter(status=DrugRecordStatus.SUBMITTED).count()
            context['approved_review_count'] = base_qs.filter(status=DrugRecordStatus.APPROVED).count()
            context['rejected_review_count'] = base_qs.filter(status=DrugRecordStatus.REJECTED).count()

    return render(request, 'core/dashboard.html', context)


# ==================== 系统管理 ====================

@login_required
def user_list(request):
    users = User.objects.all().select_related('department')
    return render(request, 'core/system/user_list.html', {'page_obj': users})


@login_required
def user_create(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        email = request.POST.get('email')
        user_type = request.POST.get('user_type')
        phone = request.POST.get('phone')
        department_id = request.POST.get('department')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, '用户名已存在')
            return redirect('user_create')
        
        department = Department.objects.get(id=department_id) if department_id else None
        User.objects.create_user(
            username=username, email=email, password=password,
            user_type=user_type, phone=phone, department=department
        )
        messages.success(request, '用户创建成功')
        return redirect('user_list')
    
    departments = Department.objects.all()
    return render(request, 'core/system/user_form.html', {'departments': departments})


@login_required
def role_list(request):
    roles = Role.objects.all()
    return render(request, 'core/system/role_list.html', {'page_obj': roles})


@login_required
def role_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        code = request.POST.get('code')
        description = request.POST.get('description')
        
        role = Role.objects.create(name=name, code=code, description=description)
        
        perm_ids = request.POST.getlist('permissions')
        role.permissions.set(Permission.objects.filter(id__in=perm_ids))
        
        messages.success(request, '角色创建成功')
        return redirect('role_list')
    
    permissions = Permission.objects.all()
    return render(request, 'core/system/role_form.html', {'permissions': permissions})


@login_required
def department_list(request):
    departments = Department.objects.all().select_related('parent', 'manager')
    return render(request, 'core/system/department_list.html', {'page_obj': departments})


@login_required
def department_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        code = request.POST.get('code')
        parent_id = request.POST.get('parent')
        manager_id = request.POST.get('manager')
        
        parent = Department.objects.get(id=parent_id) if parent_id else None
        manager = User.objects.get(id=manager_id) if manager_id else None
        
        Department.objects.create(name=name, code=code, parent=parent, manager=manager)
        messages.success(request, '部门创建成功')
        return redirect('department_list')
    
    departments = Department.objects.all()
    users = User.objects.all()
    return render(request, 'core/system/department_form.html', {'departments': departments, 'users': users})


# ==================== 客户管理 ====================

@login_required
def customer_list(request):
    customers = Customer.objects.all().select_related('owner')
    
    name = request.GET.get('name')
    status = request.GET.get('status')
    source = request.GET.get('source')
    
    if name:
        customers = customers.filter(name__icontains=name)
    if status:
        customers = customers.filter(status=status)
    if source:
        customers = customers.filter(source=source)
    
    paginator = Paginator(customers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/crm/customer_list.html', {'page_obj': page_obj})


@login_required
def customer_create(request):
    if request.method == 'POST':
        customer = Customer.objects.create(
            name=request.POST.get('name'),
            code=request.POST.get('code'),
            industry=request.POST.get('industry'),
            level=request.POST.get('level'),
            status=request.POST.get('status'),
            source=request.POST.get('source'),
            address=request.POST.get('address'),
            website=request.POST.get('website'),
            description=request.POST.get('description'),
            owner=request.user,
            created_by=request.user
        )
        
        contact_name = request.POST.get('contact_name')
        contact_phone = request.POST.get('contact_phone')
        contact_mobile = request.POST.get('contact_mobile')
        contact_email = request.POST.get('contact_email')
        if contact_name:
            Contact.objects.create(
                customer=customer, name=contact_name,
                phone=contact_phone, mobile=contact_mobile,
                email=contact_email, is_primary=True
            )
        
        messages.success(request, '客户创建成功')
        return redirect('customer_list')
    
    return render(request, 'core/crm/customer_form.html')


@login_required
def customer_detail(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    contacts = customer.contacts.all()
    follow_ups = customer.follow_ups.all()[:10]
    projects = customer.projects.all()[:10]
    return render(request, 'core/crm/customer_detail.html', {
        'customer': customer, 'contacts': contacts,
        'follow_ups': follow_ups, 'projects': projects
    })


@login_required
def customer_edit(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        customer.name = request.POST.get('name')
        customer.industry = request.POST.get('industry')
        customer.level = request.POST.get('level')
        customer.status = request.POST.get('status')
        customer.source = request.POST.get('source')
        customer.address = request.POST.get('address')
        customer.website = request.POST.get('website')
        customer.description = request.POST.get('description')
        customer.save()
        messages.success(request, '客户更新成功')
        return redirect('customer_detail', pk=pk)
    return render(request, 'core/crm/customer_form.html', {'customer': customer})


@login_required
def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    customer.delete()
    messages.success(request, '客户删除成功')
    return redirect('customer_list')


# ==================== 联系人管理 ====================

@login_required
def contact_list(request):
    contacts = Contact.objects.all().select_related('customer')
    
    name = request.GET.get('name')
    customer_id = request.GET.get('customer')
    
    if name:
        contacts = contacts.filter(name__icontains=name)
    if customer_id:
        contacts = contacts.filter(customer_id=customer_id)
    
    paginator = Paginator(contacts, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/crm/contact_list.html', {'page_obj': page_obj})


@login_required
def contact_create(request):
    if request.method == 'POST':
        customer = get_object_or_404(Customer, pk=request.POST.get('customer'))
        Contact.objects.create(
            customer=customer,
            name=request.POST.get('name'),
            position=request.POST.get('position'),
            phone=request.POST.get('phone'),
            mobile=request.POST.get('mobile'),
            email=request.POST.get('email'),
            is_primary=request.POST.get('is_primary') == 'on',
            description=request.POST.get('description')
        )
        messages.success(request, '联系人创建成功')
        return redirect('contact_list')
    
    customers = Customer.objects.all()
    return render(request, 'core/crm/contact_form.html', {'customers': customers})


@login_required
def contact_edit(request, pk):
    contact = get_object_or_404(Contact, pk=pk)
    if request.method == 'POST':
        contact.name = request.POST.get('name')
        contact.position = request.POST.get('position')
        contact.phone = request.POST.get('phone')
        contact.mobile = request.POST.get('mobile')
        contact.email = request.POST.get('email')
        contact.is_primary = request.POST.get('is_primary') == 'on'
        contact.description = request.POST.get('description')
        contact.save()
        messages.success(request, '联系人更新成功')
        return redirect('contact_list')
    customers = Customer.objects.all()
    return render(request, 'core/crm/contact_form.html', {'contact': contact, 'customers': customers})


@login_required
def contact_delete(request, pk):
    contact = get_object_or_404(Contact, pk=pk)
    contact.delete()
    messages.success(request, '联系人删除成功')
    return redirect('contact_list')


# ==================== 项目管理 ====================

@login_required
def project_list(request):
    projects = Project.objects.all().select_related('customer', 'owner')
    
    name = request.GET.get('name')
    status = request.GET.get('status')
    customer_id = request.GET.get('customer')
    
    if name:
        projects = projects.filter(name__icontains=name)
    if status:
        projects = projects.filter(status=status)
    if customer_id:
        projects = projects.filter(customer_id=customer_id)
    
    paginator = Paginator(projects, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/project/project_list.html', {'page_obj': page_obj})


@login_required
def project_create(request):
    if request.method == 'POST':
        customer_id = request.POST.get('customer')
        customer = Customer.objects.get(id=customer_id) if customer_id else None
        
        Project.objects.create(
            name=request.POST.get('name'),
            code=request.POST.get('code'),
            customer=customer,
            budget=request.POST.get('budget'),
            status=request.POST.get('status'),
            start_date=request.POST.get('start_date') or None,
            end_date=request.POST.get('end_date') or None,
            description=request.POST.get('description'),
            owner=request.user,
            created_by=request.user
        )
        messages.success(request, '项目创建成功')
        return redirect('project_list')
    
    customers = Customer.objects.all()
    return render(request, 'core/project/project_form.html', {'customers': customers})


@login_required
def project_detail(request, pk):
    project = get_object_or_404(Project, pk=pk)
    follow_ups = project.follow_ups.all()[:10]
    expenses = project.expenses.all()[:10]
    trips = project.trips.all()[:10]
    return render(request, 'core/project/project_detail.html', {
        'project': project, 'follow_ups': follow_ups,
        'expenses': expenses, 'trips': trips
    })


@login_required
def project_edit(request, pk):
    project = get_object_or_404(Project, pk=pk)
    if request.method == 'POST':
        project.name = request.POST.get('name')
        project.budget = extract_number(request.POST.get('budget', 0))
        project.status = request.POST.get('status')
        project.start_date = request.POST.get('start_date') or None
        project.end_date = request.POST.get('end_date') or None
        project.description = request.POST.get('description')
        project.save()
        messages.success(request, '项目更新成功')
        return redirect('project_detail', pk=pk)
    customers = Customer.objects.all()
    return render(request, 'core/project/project_form.html', {'project': project, 'customers': customers})


@login_required
def project_delete(request, pk):
    project = get_object_or_404(Project, pk=pk)
    project.delete()
    messages.success(request, '项目删除成功')
    return redirect('project_list')


# ==================== 跟进管理 ====================

@login_required
def followup_list(request):
    follow_ups = FollowUp.objects.all().select_related('customer', 'project', 'follower')
    
    customer_id = request.GET.get('customer')
    status = request.GET.get('status')
    follow_type = request.GET.get('follow_type')
    
    if customer_id:
        follow_ups = follow_ups.filter(customer_id=customer_id)
    if status:
        follow_ups = follow_ups.filter(status=status)
    if follow_type:
        follow_ups = follow_ups.filter(follow_type=follow_type)
    
    paginator = Paginator(follow_ups, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/followup/followup_list.html', {'page_obj': page_obj})


@login_required
def followup_create(request):
    if request.method == 'POST':
        customer = get_object_or_404(Customer, pk=request.POST.get('customer'))
        project_id = request.POST.get('project')
        project = Project.objects.get(id=project_id) if project_id else None
        
        FollowUp.objects.create(
            customer=customer,
            project=project,
            follow_type=request.POST.get('follow_type'),
            subject=request.POST.get('subject'),
            content=request.POST.get('content'),
            status=request.POST.get('status'),
            next_action=request.POST.get('next_action'),
            next_date=request.POST.get('next_date') or None,
            follower=request.user
        )
        messages.success(request, '跟进记录创建成功')
        return redirect('followup_list')
    
    customers = Customer.objects.all()
    projects = Project.objects.all()
    return render(request, 'core/followup/followup_form.html', {'customers': customers, 'projects': projects})


@login_required
def followup_edit(request, pk):
    followup = get_object_or_404(FollowUp, pk=pk)
    if request.method == 'POST':
        followup.subject = request.POST.get('subject')
        followup.content = request.POST.get('content')
        followup.status = request.POST.get('status')
        followup.next_action = request.POST.get('next_action')
        followup.next_date = request.POST.get('next_date') or None
        followup.save()
        messages.success(request, '跟进记录更新成功')
        return redirect('followup_list')
    customers = Customer.objects.all()
    projects = Project.objects.all()
    return render(request, 'core/followup/followup_form.html', {'followup': followup, 'customers': customers, 'projects': projects})


@login_required
def followup_delete(request, pk):
    followup = get_object_or_404(FollowUp, pk=pk)
    followup.delete()
    messages.success(request, '跟进记录删除成功')
    return redirect('followup_list')


# ==================== 出差管理 ====================

@login_required
def trip_list(request):
    trips = Trip.objects.all().select_related('user', 'project')
    
    user_id = request.GET.get('user')
    status = request.GET.get('status')
    
    if user_id:
        trips = trips.filter(user_id=user_id)
    if status:
        trips = trips.filter(status=status)
    
    paginator = Paginator(trips, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/trip/trip_list.html', {'page_obj': page_obj})


@login_required
def trip_create(request):
    if request.method == 'POST':
        project_id = request.POST.get('project')
        project = Project.objects.get(id=project_id) if project_id else None
        
        Trip.objects.create(
            user=request.user,
            project=project,
            destination=request.POST.get('destination'),
            purpose=request.POST.get('purpose'),
            start_date=request.POST.get('start_date'),
            end_date=request.POST.get('end_date'),
            status='submitted',
            estimated_cost=request.POST.get('estimated_cost') or 0,
            notes=request.POST.get('notes')
        )
        messages.success(request, '出差申请提交成功')
        return redirect('trip_list')
    
    projects = Project.objects.all()
    return render(request, 'core/trip/trip_form.html', {'projects': projects})


@login_required
def trip_approve(request, pk):
    trip = get_object_or_404(Trip, pk=pk)
    action = request.POST.get('action')
    
    if action == 'approve':
        trip.status = 'approved'
    elif action == 'reject':
        trip.status = 'rejected'
    
    trip.save()
    messages.success(request, '审批完成')
    return redirect('trip_list')


# ==================== 业务费管理 ====================

@login_required
def expense_list(request):
    expenses = Expense.objects.all().select_related('user', 'project', 'trip')
    
    user_id = request.GET.get('user')
    status = request.GET.get('status')
    expense_type = request.GET.get('expense_type')
    
    if user_id:
        expenses = expenses.filter(user_id=user_id)
    if status:
        expenses = expenses.filter(status=status)
    if expense_type:
        expenses = expenses.filter(expense_type=expense_type)
    
    paginator = Paginator(expenses, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/expense/expense_list.html', {'page_obj': page_obj})


@login_required
def expense_create(request):
    if request.method == 'POST':
        project_id = request.POST.get('project')
        trip_id = request.POST.get('trip')
        
        project = Project.objects.get(id=project_id) if project_id else None
        trip = Trip.objects.get(id=trip_id) if trip_id else None
        
        Expense.objects.create(
            user=request.user,
            project=project,
            trip=trip,
            expense_type=request.POST.get('expense_type'),
            amount=request.POST.get('amount'),
            expense_date=request.POST.get('expense_date'),
            description=request.POST.get('description'),
            status='submitted'
        )
        messages.success(request, '费用申请提交成功')
        return redirect('expense_list')
    
    projects = Project.objects.all()
    trips = Trip.objects.filter(user=request.user)
    return render(request, 'core/expense/expense_form.html', {'projects': projects, 'trips': trips})


@login_required
def expense_approve(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    action = request.POST.get('action')
    
    if action == 'approve':
        expense.status = 'approved'
    elif action == 'reject':
        expense.status = 'rejected'
    
    expense.save()
    messages.success(request, '审批完成')
    return redirect('expense_list')


# ==================== 药品差比价管理 ====================

def export_drugs_excel(request):
    code = request.GET.get('code')
    generic_name = request.GET.get('generic_name')

    drugs = Drug.objects.all().order_by('id')
    if code:
        drugs = drugs.filter(code__icontains=code)
    if generic_name:
        drugs = drugs.filter(generic_name__icontains=generic_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "药品数据"

    headers = [
        '统编代码', '药品分类', '通用名', '目录剂型', '表述剂型', '规格包装',
        '标化持有人', '目录序号&标化持有人', '含量', '含量单位', '装量', '装量单位',
        '计价数量单位', '包装数量', '计价数量', '服用天数', '挂网价', '特殊标注',
        '标准品标记', '差比价', '价差对比', '单体差比值',
    ]
    ws.append(headers)

    for drug in drugs:
        ws.append([
            drug.code, drug.drug_category, drug.generic_name,
            drug.catalog_dosage_form, drug.description_dosage_form, drug.spec_package,
            drug.standard_holder, drug.catalog_holder,
            drug.content, drug.content_unit, drug.volume, drug.volume_unit,
            drug.quantity_unit, drug.package_quantity, drug.quantity, drug.usage_days,
            float(drug.network_price) if drug.network_price else 0,
            drug.special_note,
            drug.standard_mark,
            float(drug.standard_price) if drug.standard_price else 0,
            float(drug.price_diff) if drug.price_diff else 0,
            float(drug.unit_value) if drug.unit_value else 0,
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=药品数据_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
def drug_list(request):
    drugs = Drug.objects.all().order_by('id')
    
    code = request.GET.get('code')
    generic_name = request.GET.get('generic_name')
    
    if code:
        drugs = drugs.filter(code__icontains=code)
    if generic_name:
        drugs = drugs.filter(generic_name__icontains=generic_name)
    
    paginator = Paginator(drugs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/drug/drug_list.html', {'page_obj': page_obj})


@login_required
def drug_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, '请选择文件')
            return redirect('drug_import')

        try:
            df = pd.read_excel(excel_file)

            for _, row in df.iterrows():
                defaults = {
                    'drug_category': _safe_str(row.get('药品分类', '')),
                    'generic_name': _safe_str(row.get('通用名', '')),
                    'catalog_dosage_form': _safe_str(row.get('目录剂型', '')),
                    'description_dosage_form': _safe_str(row.get('表述剂型', '')),
                    'spec_package': _safe_str(row.get('规格包装', '')),
                    'content': _safe_str(row.get('含量', '')),
                    'content_unit': _safe_str(row.get('含量单位', '')),
                    'volume': _safe_str(row.get('装量', '')),
                    'volume_unit': _safe_str(row.get('装量单位', '')),
                    'quantity': _safe_str(row.get('计价数量', '')),
                    'quantity_unit': _safe_str(row.get('计价数量单位', '')),
                    'package_quantity': _safe_str(row.get('包装数量', '')),
                    'usage_days': _safe_str(row.get('服用天数', '')),
                    'standard_holder': _safe_str(row.get('标化持有人', '')),
                    'catalog_holder': _safe_str(row.get('目录序号&标化持有人', '')),
                    'catalog_name': _safe_str(row.get('医保目录名', '')),
                    'special_note': _safe_str(row.get('特殊标注', '')),
                    'network_price': extract_number(row.get('挂网价', 0)),
                }
                Drug.objects.update_or_create(
                    code=str(row.get('统编代码', '')),
                    defaults=defaults,
                )

            messages.success(request, f'成功导入 {len(df)} 条药品数据')
        except Exception as e:
            messages.error(request, f'导入失败: {str(e)}')

        return redirect('drug_list')

    return render(request, 'core/drug/drug_import.html')


@login_required
def drug_calculate(request):
    if request.method == 'POST':
        match_catalog = request.POST.get('match_catalog') == 'on'
        match_generic = request.POST.get('match_generic') == 'on'

        if not match_catalog and not match_generic:
            messages.error(request, '请至少选择一种匹配方式')
            return redirect('drug_calculate')

        drugs = Drug.objects.all()
        drug_list = []
        for drug in drugs:
            drug_list.append({
                'id': drug.id,
                'code': drug.code,
                'drug_category': drug.drug_category,
                'generic_name': drug.generic_name,
                'catalog_dosage_form': drug.catalog_dosage_form,
                'description_dosage_form': drug.description_dosage_form,
                'content': drug.content,
                'volume': drug.volume,
                'quantity': drug.quantity,
                'usage_days': drug.usage_days,
                'standard_holder': drug.standard_holder,
                'catalog_name': drug.catalog_name,
                'special_note': drug.special_note,
                'spec_package': drug.spec_package,
                'network_price': drug.network_price,
            })

        drug_list = process_all_drugs(
            drug_list,
            match_catalog=match_catalog,
            match_generic=match_generic,
        )

        for drug_data in drug_list:
            Drug.objects.filter(id=drug_data['id']).update(
                is_standard=drug_data.get('is_standard', False),
                standard_mark=drug_data.get('standard_mark', ''),
                standard_price=drug_data.get('standard_price', 0),
                price_diff=drug_data.get('price_diff', 0),
                unit_value=drug_data.get('unit_value', 0),
            )

        mode_desc = []
        if match_catalog:
            mode_desc.append('同医保目录')
        if match_generic:
            mode_desc.append('同通用名')
        messages.success(request, f'差比价计算完成（匹配方式：{"+".join(mode_desc)}）')
        return redirect('drug_list')

    return render(request, 'core/drug/drug_calculate.html')


@login_required
def pharmacy_list(request):
    pharmacies = Pharmacy.objects.all()
    return render(request, 'core/drug/pharmacy_list.html', {'page_obj': pharmacies})


@login_required
def pharmacy_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        df = pd.read_excel(excel_file)
        
        for _, row in df.iterrows():
            Pharmacy.objects.update_or_create(
                pharmacy_code=str(row.get('药店编码', '')),
                defaults={
                    'pharmacy_name': str(row.get('药店名称', '')),
                    'medical_insurance_code': str(row.get('医保编码', '')),
                    'district': str(row.get('区县', '')),
                    'address': str(row.get('地址', '')),
                }
            )
        
        messages.success(request, f'成功导入 {len(df)} 条药店数据')
        return redirect('pharmacy_list')
    
    return render(request, 'core/drug/pharmacy_import.html')


@login_required
def company_list(request):
    companies = PharmaceuticalCompany.objects.all()
    return render(request, 'core/drug/company_list.html', {'page_obj': companies})


@login_required
def company_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        df = pd.read_excel(excel_file)
        
        for _, row in df.iterrows():
            PharmaceuticalCompany.objects.update_or_create(
                company_code=str(row.get('机构编码', '')),
                defaults={
                    'company_name': str(row.get('企业名称', '')),
                    'contact_person': str(row.get('联系人', '')),
                    'phone': str(row.get('电话', '')),
                }
            )
        
        messages.success(request, f'成功导入 {len(df)} 条药企数据')
        return redirect('company_list')
    
    return render(request, 'core/drug/company_import.html')


# ==================== 药事所监督科 ====================

@login_required
def supervisor_pharmacy_record_list(request):
    """药事所监督科 - 药品备案价查询：查看所有药店的备案价记录"""
    records = PharmacyRecord.objects.all().select_related('drug', 'pharmacy').order_by('-created_at')

    drug_code = request.GET.get('drug_code')
    generic_name = request.GET.get('generic_name')
    pharmacy_name = request.GET.get('pharmacy_name')
    district = request.GET.get('district')
    status = request.GET.get('status')
    record_date = request.GET.get('record_date')

    if drug_code:
        records = records.filter(drug__code__icontains=drug_code)
    if generic_name:
        records = records.filter(drug__generic_name__icontains=generic_name)
    if pharmacy_name:
        records = records.filter(pharmacy__pharmacy_name__icontains=pharmacy_name)
    if district:
        records = records.filter(pharmacy__district__icontains=district)
    if status:
        records = records.filter(status=status)
    if record_date:
        records = records.filter(record_date__startswith=record_date)

    paginator = Paginator(records, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'core/yaoshisuo/pharmacy_record_list.html', {'page_obj': page_obj})


def export_supervisor_pharmacy_records_excel(request):
    """药事所监督科 - 导出药店备案价"""
    records = PharmacyRecord.objects.all().select_related('drug', 'pharmacy')

    drug_code = request.GET.get('drug_code')
    generic_name = request.GET.get('generic_name')
    pharmacy_name = request.GET.get('pharmacy_name')
    district = request.GET.get('district')
    status = request.GET.get('status')
    record_date = request.GET.get('record_date')

    if drug_code:
        records = records.filter(drug__code__icontains=drug_code)
    if generic_name:
        records = records.filter(drug__generic_name__icontains=generic_name)
    if pharmacy_name:
        records = records.filter(pharmacy__pharmacy_name__icontains=pharmacy_name)
    if district:
        records = records.filter(pharmacy__district__icontains=district)
    if status:
        records = records.filter(status=status)
    if record_date:
        records = records.filter(record_date__startswith=record_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "药品备案价查询"

    headers = ['序号', '区县', '药店名称', '药店编码', '医保编码', '统编代码', '通用名',
               '上市许可持有人', '规格包装', '挂网价', '备案价', '数据年月', '状态', '提交时间']
    ws.append(headers)

    for idx, record in enumerate(records, 1):
        ws.append([
            idx,
            record.pharmacy.district,
            record.pharmacy.pharmacy_name,
            record.pharmacy.pharmacy_code,
            record.pharmacy.medical_insurance_code,
            record.drug.code,
            record.drug.generic_name,
            record.drug.standard_holder,
            record.drug.spec_package,
            float(record.drug.network_price) if record.drug.network_price else 0,
            float(record.record_price),
            str(record.record_date),
            record.get_status_display(),
            record.created_at.strftime('%Y-%m-%d %H:%M') if record.created_at else '',
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=药品备案价_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response

@login_required
def supervisor_drug_list(request):
    drugs = Drug.objects.all().order_by('id')
    
    code = request.GET.get('code')
    generic_name = request.GET.get('generic_name')
    
    if code:
        drugs = drugs.filter(code__icontains=code)
    if generic_name:
        drugs = drugs.filter(generic_name__icontains=generic_name)
    
    paginator = Paginator(drugs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/yaoshisuo/drug_list.html', {'page_obj': page_obj})


@login_required
def supervisor_pharmacy_list(request):
    pharmacies = Pharmacy.objects.all()
    
    name = request.GET.get('name')
    district = request.GET.get('district')
    
    if name:
        pharmacies = pharmacies.filter(pharmacy_name__icontains=name)
    if district:
        pharmacies = pharmacies.filter(district__icontains=district)
    
    paginator = Paginator(pharmacies, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/yaoshisuo/pharmacy_list.html', {'page_obj': page_obj})


@login_required
def supervisor_pharmacy_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        df = pd.read_excel(excel_file)
        
        for _, row in df.iterrows():
            Pharmacy.objects.update_or_create(
                pharmacy_code=str(row.get('药店编码', '')),
                defaults={
                    'pharmacy_name': str(row.get('药店名称', '')),
                    'medical_insurance_code': str(row.get('医保编码', '')),
                    'district': str(row.get('区县', '')),
                    'address': str(row.get('地址', '')),
                    'phone': str(row.get('电话', '')),
                }
            )
        
        messages.success(request, f'成功导入 {len(df)} 条药店数据')
        return redirect('supervisor_pharmacy_list')
    
    return render(request, 'core/yaoshisuo/pharmacy_import.html')


@login_required
def supervisor_pharmacy_create(request):
    if request.method == 'POST':
        Pharmacy.objects.create(
            pharmacy_code=request.POST.get('pharmacy_code'),
            pharmacy_name=request.POST.get('pharmacy_name'),
            medical_insurance_code=request.POST.get('medical_insurance_code'),
            district=request.POST.get('district'),
            address=request.POST.get('address'),
            phone=request.POST.get('phone'),
        )
        messages.success(request, '药店创建成功')
        return redirect('supervisor_pharmacy_list')
    
    return render(request, 'core/yaoshisuo/pharmacy_form.html')


@login_required
def supervisor_pharmacy_delete(request, pk):
    pharmacy = get_object_or_404(Pharmacy, pk=pk)
    pharmacy.delete()
    messages.success(request, '药店删除成功')
    return redirect('supervisor_pharmacy_list')


@login_required
def supervisor_pharmacy_link_user(request, pk):
    """药事所监督科 - 药店编码关系维护：关联药店与用户账号"""
    pharmacy = get_object_or_404(Pharmacy, pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action')
        user_id = request.POST.get('user_id')

        if action == 'link' and user_id:
            user = get_object_or_404(User, pk=user_id)
            # 更新用户类型为药店
            user.user_type = 'pharmacy'
            user.organization = pharmacy.pharmacy_name
            user.medical_insurance_code = pharmacy.medical_insurance_code
            user.save()
            # 创建或更新药店用户关联
            PharmacyUser.objects.update_or_create(
                user=user,
                defaults={'pharmacy': pharmacy},
            )
            messages.success(request, f'已将用户"{user.username}"关联至药店"{pharmacy.pharmacy_name}"')

        elif action == 'unlink' and user_id:
            user = get_object_or_404(User, pk=user_id)
            try:
                pu = PharmacyUser.objects.get(user=user, pharmacy=pharmacy)
                pu.delete()
                messages.success(request, f'已解除用户"{user.username}"与药店"{pharmacy.pharmacy_name}"的关联')
            except PharmacyUser.DoesNotExist:
                messages.warning(request, '该用户未关联此药店')

        elif action == 'create_user':
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '')
            if username and password:
                if User.objects.filter(username=username).exists():
                    messages.error(request, f'用户名"{username}"已存在')
                else:
                    user = User.objects.create_user(
                        username=username,
                        password=password,
                        user_type='pharmacy',
                        organization=pharmacy.pharmacy_name,
                        medical_insurance_code=pharmacy.medical_insurance_code,
                    )
                    PharmacyUser.objects.create(user=user, pharmacy=pharmacy)
                    messages.success(request, f'已创建药店用户"{username}"并关联至药店"{pharmacy.pharmacy_name}"')

        return redirect('supervisor_pharmacy_link_user', pk=pk)

    # 获取已关联的用户
    linked_users = PharmacyUser.objects.filter(pharmacy=pharmacy).select_related('user')
    # 获取可选用户（药店类型且未关联此药店）
    available_users = User.objects.filter(user_type='pharmacy').exclude(
        pk__in=PharmacyUser.objects.filter(pharmacy=pharmacy).values('user_id')
    )

    return render(request, 'core/yaoshisuo/pharmacy_link_user.html', {
        'pharmacy': pharmacy,
        'linked_users': linked_users,
        'available_users': available_users,
    })


@login_required
def supervisor_anchor_price_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        record_date = request.POST.get('record_date')
        if record_date and len(record_date) == 7:
            record_date = record_date + '-01'

        df = pd.read_excel(excel_file)

        for _, row in df.iterrows():
            drug_code = str(row.get('统编代码', ''))
            drug = Drug.objects.filter(code=drug_code).first()
            if drug:
                AnchorPrice.objects.update_or_create(
                    drug=drug,
                    record_date=record_date,
                    defaults={
                        'anchor_price': extract_number(row.get('锚点价格', 0)),
                        'adjust_ratio': extract_number(row.get('调整倍率', 1.0)),
                        'created_by': request.user,
                    }
                )
        
        messages.success(request, f'成功导入锚点价格')
        return redirect('supervisor_anchor_price_list')
    
    return render(request, 'core/yaoshisuo/import_anchor_price.html')


@login_required
def supervisor_anchor_price_list(request):
    anchors = AnchorPrice.objects.all().select_related('drug').order_by('id')
    
    drug_code = request.GET.get('drug_code')
    generic_name = request.GET.get('generic_name')
    
    if drug_code:
        anchors = anchors.filter(drug__code__icontains=drug_code)
    if generic_name:
        anchors = anchors.filter(drug__generic_name__icontains=generic_name)
    
    paginator = Paginator(anchors, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/yaoshisuo/anchor_price_list.html', {'page_obj': page_obj})


def export_anchor_price_excel(request):
    drug_code = request.GET.get('drug_code')
    generic_name = request.GET.get('generic_name')
    
    anchors = AnchorPrice.objects.all().select_related('drug')
    
    if drug_code:
        anchors = anchors.filter(drug__code__icontains=drug_code)
    if generic_name:
        anchors = anchors.filter(drug__generic_name__icontains=generic_name)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "锚点价格"
    
    headers = ['统编代码', '通用名', '锚点价格', '调整倍率', '治理价格', '数据年月']
    ws.append(headers)
    
    for anchor in anchors:
        ws.append([
            anchor.drug.code,
            anchor.drug.generic_name,
            float(anchor.anchor_price),
            float(anchor.adjust_ratio),
            float(anchor.target_price),
            str(anchor.record_date)
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=锚点价格_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
def supervisor_anchor_price_edit(request, pk):
    anchor = get_object_or_404(AnchorPrice, pk=pk)
    if request.method == 'POST':
        anchor.anchor_price = extract_number(request.POST.get('anchor_price', 0))
        anchor.adjust_ratio = extract_number(request.POST.get('adjust_ratio', 1.0))
        anchor.save()
        messages.success(request, '锚点价格更新成功')
        return redirect('supervisor_anchor_price_list')
    return render(request, 'core/yaoshisuo/anchor_price_form.html', {'anchor': anchor})


@login_required
def supervisor_company_list(request):
    companies = PharmaceuticalCompany.objects.all().order_by('id')
    
    name = request.GET.get('name')
    
    if name:
        companies = companies.filter(company_name__icontains=name)
    
    paginator = Paginator(companies, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/yaoshisuo/company_list.html', {'page_obj': page_obj})


@login_required
def supervisor_company_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        df = pd.read_excel(excel_file)
        
        for _, row in df.iterrows():
            PharmaceuticalCompany.objects.update_or_create(
                company_code=str(row.get('机构编码', '')),
                defaults={
                    'company_name': str(row.get('企业名称', '')),
                    'contact_person': str(row.get('联系人', '')),
                    'phone': str(row.get('电话', '')),
                    'address': str(row.get('地址', '')),
                }
            )
        
        messages.success(request, f'成功导入 {len(df)} 条药企数据')
        return redirect('supervisor_company_list')
    
    return render(request, 'core/yaoshisuo/company_import.html')


@login_required
def supervisor_company_create(request):
    if request.method == 'POST':
        PharmaceuticalCompany.objects.create(
            company_code=request.POST.get('company_code'),
            company_name=request.POST.get('company_name'),
            contact_person=request.POST.get('contact_person'),
            phone=request.POST.get('phone'),
            address=request.POST.get('address'),
        )
        messages.success(request, '药企创建成功')
        return redirect('supervisor_company_list')
    
    return render(request, 'core/yaoshisuo/company_form.html')


@login_required
def supervisor_company_delete(request, pk):
    company = get_object_or_404(PharmaceuticalCompany, pk=pk)
    company.delete()
    messages.success(request, '药企删除成功')
    return redirect('supervisor_company_list')


@login_required
def supervisor_company_link_user(request, pk):
    """药事所监督科 - 药企编码关系维护：关联药企与用户账号"""
    company = get_object_or_404(PharmaceuticalCompany, pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action')
        user_id = request.POST.get('user_id')

        if action == 'link' and user_id:
            user = get_object_or_404(User, pk=user_id)
            user.user_type = 'company'
            user.organization = company.company_name
            user.save()
            CompanyUser.objects.update_or_create(
                user=user,
                defaults={'company': company},
            )
            messages.success(request, f'已将用户"{user.username}"关联至药企"{company.company_name}"')

        elif action == 'unlink' and user_id:
            user = get_object_or_404(User, pk=user_id)
            try:
                cu = CompanyUser.objects.get(user=user, company=company)
                cu.delete()
                messages.success(request, f'已解除用户"{user.username}"与药企"{company.company_name}"的关联')
            except CompanyUser.DoesNotExist:
                messages.warning(request, '该用户未关联此药企')

        elif action == 'create_user':
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '')
            if username and password:
                if User.objects.filter(username=username).exists():
                    messages.error(request, f'用户名"{username}"已存在')
                else:
                    user = User.objects.create_user(
                        username=username,
                        password=password,
                        user_type='company',
                        organization=company.company_name,
                    )
                    CompanyUser.objects.create(user=user, company=company)
                    messages.success(request, f'已创建药企用户"{username}"并关联至药企"{company.company_name}"')

        return redirect('supervisor_company_link_user', pk=pk)

    linked_users = CompanyUser.objects.filter(company=company).select_related('user')
    available_users = User.objects.filter(user_type='company').exclude(
        pk__in=CompanyUser.objects.filter(company=company).values('user_id')
    )

    return render(request, 'core/yaoshisuo/company_link_user.html', {
        'company': company,
        'linked_users': linked_users,
        'available_users': available_users,
    })


# ==================== 药事所药品科 ====================

@login_required
def drug_section_drug_list(request):
    drugs = Drug.objects.all()
    
    code = request.GET.get('code')
    generic_name = request.GET.get('generic_name')
    
    if code:
        drugs = drugs.filter(code__icontains=code)
    if generic_name:
        drugs = drugs.filter(generic_name__icontains=generic_name)
    
    paginator = Paginator(drugs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/yaoqi/drug_list.html', {'page_obj': page_obj})


@login_required
def drug_section_review_list(request):
    """药事所药品科 - 药品挂网价审核：审核药企提交的申报价"""
    reviews = DrugPriceReview.objects.all().select_related('drug', 'reviewer').order_by('-created_at')

    drug_code = request.GET.get('drug_code')
    generic_name = request.GET.get('generic_name')
    status = request.GET.get('status')

    if drug_code:
        reviews = reviews.filter(drug__code__icontains=drug_code)
    if generic_name:
        reviews = reviews.filter(drug__generic_name__icontains=generic_name)
    if status:
        reviews = reviews.filter(status=status)
    else:
        # 默认显示待整改的
        reviews = reviews.filter(status=DrugRecordStatus.PENDING)

    paginator = Paginator(reviews, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'core/yaoqi/review_list.html', {'page_obj': page_obj})


@login_required
def drug_section_review_approve(request, pk):
    review = get_object_or_404(DrugPriceReview, pk=pk)
    action = request.POST.get('action')

    if action == 'approve':
        review.status = DrugRecordStatus.APPROVED
        # 更新药品挂网价为审核通过的整改价
        review.drug.network_price = review.proposed_price
        review.drug.save()
        # 联动更新：将该药品相关的药企申报记录状态也更新为审核通过
        CompanyRecord.objects.filter(
            drug=review.drug,
            declared_price=review.proposed_price,
            status=DrugRecordStatus.SUBMITTED,
        ).update(status=DrugRecordStatus.APPROVED)
    elif action == 'reject':
        review.status = DrugRecordStatus.REJECTED
        # 联动更新：将该药品相关的药企申报记录状态也更新为审核不通过
        CompanyRecord.objects.filter(
            drug=review.drug,
            declared_price=review.proposed_price,
            status=DrugRecordStatus.SUBMITTED,
        ).update(status=DrugRecordStatus.REJECTED)

    review.reviewer = request.user
    review.reviewed_at = datetime.now()
    review.save()

    messages.success(request, '审核完成')
    return redirect('drug_section_review_list')


@login_required
def drug_section_company_list(request):
    companies = PharmaceuticalCompany.objects.all().order_by('id')
    
    name = request.GET.get('name')
    
    if name:
        companies = companies.filter(company_name__icontains=name)
    
    paginator = Paginator(companies, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/yaoqi/company_list.html', {'page_obj': page_obj})


@login_required
def drug_section_company_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        df = pd.read_excel(excel_file)

        for _, row in df.iterrows():
            PharmaceuticalCompany.objects.update_or_create(
                company_code=str(row.get('机构编码', '')),
                defaults={
                    'company_name': str(row.get('企业名称', '')),
                    'contact_person': str(row.get('联系人', '')),
                    'phone': str(row.get('电话', '')),
                }
            )

        messages.success(request, f'成功导入 {len(df)} 条药企数据')
        return redirect('drug_section_company_list')

    return render(request, 'core/yaoqi/company_import.html')


@login_required
def drug_section_company_create(request):
    if request.method == 'POST':
        PharmaceuticalCompany.objects.create(
            company_code=request.POST.get('company_code'),
            company_name=request.POST.get('company_name'),
            contact_person=request.POST.get('contact_person'),
            phone=request.POST.get('phone'),
            address=request.POST.get('address'),
        )
        messages.success(request, '药企创建成功')
        return redirect('drug_section_company_list')
    return render(request, 'core/yaoshisuo/company_form.html')


@login_required
def drug_section_company_delete(request, pk):
    company = get_object_or_404(PharmaceuticalCompany, pk=pk)
    company.delete()
    messages.success(request, '药企删除成功')
    return redirect('drug_section_company_list')


@login_required
def drug_section_company_link_user(request, pk):
    """药事所药品科 - 药企编码关系维护：关联药企与用户账号"""
    company = get_object_or_404(PharmaceuticalCompany, pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action')
        user_id = request.POST.get('user_id')

        if action == 'link' and user_id:
            user = get_object_or_404(User, pk=user_id)
            user.user_type = 'company'
            user.organization = company.company_name
            user.save()
            CompanyUser.objects.update_or_create(
                user=user,
                defaults={'company': company},
            )
            messages.success(request, f'已将用户"{user.username}"关联至药企"{company.company_name}"')

        elif action == 'unlink' and user_id:
            user = get_object_or_404(User, pk=user_id)
            try:
                cu = CompanyUser.objects.get(user=user, company=company)
                cu.delete()
                messages.success(request, f'已解除用户"{user.username}"与药企"{company.company_name}"的关联')
            except CompanyUser.DoesNotExist:
                messages.warning(request, '该用户未关联此药企')

        elif action == 'create_user':
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '')
            if username and password:
                if User.objects.filter(username=username).exists():
                    messages.error(request, f'用户名"{username}"已存在')
                else:
                    user = User.objects.create_user(
                        username=username,
                        password=password,
                        user_type='company',
                        organization=company.company_name,
                    )
                    CompanyUser.objects.create(user=user, company=company)
                    messages.success(request, f'已创建药企用户"{username}"并关联至药企"{company.company_name}"')

        return redirect('drug_section_company_link_user', pk=pk)

    linked_users = CompanyUser.objects.filter(company=company).select_related('user')
    available_users = User.objects.filter(user_type='company').exclude(
        pk__in=CompanyUser.objects.filter(company=company).values('user_id')
    )

    return render(request, 'core/yaoqi/company_link_user.html', {
        'company': company,
        'linked_users': linked_users,
        'available_users': available_users,
    })


# ==================== 药店端 ====================

def export_pharmacy_records_excel(request):
    pharmacy = None
    if request.user.user_type == 'pharmacy':
        try:
            pharmacy = request.user.pharmacy_profile.pharmacy
        except:
            pharmacy = None
    
    if pharmacy:
        records = PharmacyRecord.objects.filter(pharmacy=pharmacy).select_related('drug', 'pharmacy')
    else:
        records = PharmacyRecord.objects.none()
    
    drug_code = request.GET.get('drug_code')
    status = request.GET.get('status')
    record_date = request.GET.get('record_date')
    
    if drug_code:
        records = records.filter(drug__code__icontains=drug_code)
    if status:
        records = records.filter(status=status)
    if record_date:
        records = records.filter(record_date=record_date)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "药店备案价"
    
    headers = ['序号', '注册区县', '药店名称', '药店代码', '统编代码', '通用名', '上市许可持有人', '规格包装', '备案价', '数据年月', '状态']
    ws.append(headers)
    
    for idx, record in enumerate(records, 1):
        ws.append([
            idx,
            record.pharmacy.district,
            record.pharmacy.pharmacy_name,
            record.pharmacy.pharmacy_code,
            record.drug.code,
            record.drug.generic_name,
            record.drug.standard_holder,
            record.drug.spec_package,
            float(record.record_price),
            str(record.record_date),
            record.get_status_display()
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=药店备案价_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
def pharmacy_drug_list(request):
    drugs = Drug.objects.all()
    
    code = request.GET.get('code')
    generic_name = request.GET.get('generic_name')
    
    if code:
        drugs = drugs.filter(code__icontains=code)
    if generic_name:
        drugs = drugs.filter(generic_name__icontains=generic_name)
    
    paginator = Paginator(drugs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/yaodian/drug_list.html', {'page_obj': page_obj})


@login_required
def pharmacy_record_submit(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        record_date = request.POST.get('record_date', '').strip()

        if not excel_file:
            messages.error(request, '请选择要上传的Excel文件')
            return redirect('pharmacy_record_submit')

        if not record_date:
            messages.error(request, '请选择数据年月')
            return redirect('pharmacy_record_submit')

        # 处理年月格式：YYYY-MM → YYYY-MM-DD (取当月第一天)
        if len(record_date) == 7:
            record_date = record_date + '-01'
        elif len(record_date) == 10:
            pass  # 已经是 YYYY-MM-DD 格式
        else:
            messages.error(request, f'日期格式不正确：{record_date}')
            return redirect('pharmacy_record_submit')

        # 获取当前药店用户关联的药店
        pharmacy = None
        if request.user.user_type == 'pharmacy':
            try:
                pharmacy = request.user.pharmacy_profile.pharmacy
            except PharmacyUser.DoesNotExist:
                pharmacy = None

        if not pharmacy:
            messages.error(request, '您的账号尚未关联药店，请联系药事所监督科进行药店编码关系维护。')
            return redirect('pharmacy_record_submit')

        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            messages.error(request, f'Excel文件读取失败：{str(e)}，请确认文件格式正确')
            return redirect('pharmacy_record_submit')

        if len(df) == 0:
            messages.error(request, 'Excel文件中没有数据')
            return redirect('pharmacy_record_submit')

        success_count = 0
        skip_count = 0
        error_messages = []
        current_pharmacy_name = pharmacy.pharmacy_name

        for idx, row in df.iterrows():
            row_num = idx + 2  # Excel行号（第1行是表头）
            drug_code = str(row.get('统编代码', '')).strip()
            record_price = extract_number(row.get('备案价', 0))

            # 校验药店名称：如果Excel中包含药店名称列，则必须与当前药店名称一致
            excel_pharmacy_name = row.get('药店名称', '')
            if excel_pharmacy_name and str(excel_pharmacy_name).strip():
                excel_name = str(excel_pharmacy_name).strip()
                if excel_name != current_pharmacy_name:
                    error_messages.append(
                        f'第{row_num}行：药店名称"{excel_name}"与当前登录药店"{current_pharmacy_name}"不一致，'
                        f'只能提交本药店的备案价数据'
                    )
                    skip_count += 1
                    continue

            if not drug_code:
                error_messages.append(f'第{row_num}行：统编代码不能为空')
                skip_count += 1
                continue

            drug = Drug.objects.filter(code=drug_code).first()
            if not drug:
                error_messages.append(f'第{row_num}行：统编代码"{drug_code}"对应的药品不存在')
                skip_count += 1
                continue

            if record_price <= 0:
                error_messages.append(f'第{row_num}行：备案价必须大于0，当前值为{record_price}')
                skip_count += 1
                continue

            # 价格校验：备案价需在挂网价的合理范围内
            network_price = drug.network_price
            if network_price > 0:
                max_price = round(network_price * Decimal('1.15'), 2)
                min_price = round(network_price / Decimal('1.3'), 2)

                if record_price > max_price:
                    error_messages.append(
                        f'第{row_num}行：备案价{record_price}元高于挂网价1.15倍上限{max_price}元 '
                        f'(药品：{drug.generic_name}，挂网价：{network_price}元)'
                    )
                    skip_count += 1
                    continue
                if record_price < min_price:
                    error_messages.append(
                        f'第{row_num}行：备案价{record_price}元低于挂网价1/1.3倍下限{min_price}元 '
                        f'(药品：{drug.generic_name}，挂网价：{network_price}元)'
                    )
                    skip_count += 1
                    continue

            # 创建或更新备案价记录
            PharmacyRecord.objects.update_or_create(
                pharmacy=pharmacy,
                drug=drug,
                record_date=record_date,
                defaults={
                    'record_price': record_price,
                    'status': DrugRecordStatus.SUBMITTED,
                }
            )
            success_count += 1

        # 汇总反馈
        if success_count > 0:
            messages.success(request, f'成功提交 {success_count} 条备案价记录')
        if skip_count > 0:
            messages.warning(request, f'跳过 {skip_count} 条记录（详见下方提示）')
        if error_messages:
            for msg in error_messages[:20]:
                messages.warning(request, msg)
            if len(error_messages) > 20:
                messages.warning(request, f'...还有 {len(error_messages) - 20} 条错误未显示')
        if success_count == 0 and skip_count == 0:
            messages.error(request, '没有提交任何数据，请检查Excel文件内容')

        return redirect('pharmacy_record_list')

    return render(request, 'core/yaodian/record_submit.html')


@login_required
def pharmacy_record_list(request):
    pharmacy = None
    if request.user.user_type == 'pharmacy':
        try:
            pharmacy = request.user.pharmacy_profile.pharmacy
        except:
            pharmacy = None
    
    if pharmacy:
        records = PharmacyRecord.objects.filter(pharmacy=pharmacy).select_related('drug', 'pharmacy')
    else:
        records = PharmacyRecord.objects.none()
    
    drug_code = request.GET.get('drug_code')
    generic_name = request.GET.get('generic_name')
    spec_package = request.GET.get('spec_package')
    status = request.GET.get('status')
    record_date = request.GET.get('record_date')
    
    if drug_code:
        records = records.filter(drug__code__icontains=drug_code)
    if generic_name:
        records = records.filter(drug__generic_name__icontains=generic_name)
    if spec_package:
        records = records.filter(drug__spec_package__icontains=spec_package)
    if status:
        records = records.filter(status=status)
    if record_date:
        records = records.filter(record_date=record_date)
    
    paginator = Paginator(records, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/yaodian/record_list.html', {'page_obj': page_obj})


# ==================== 药企端 ====================

@login_required
def company_drug_list(request):
    drugs = Drug.objects.all()
    
    code = request.GET.get('code')
    generic_name = request.GET.get('generic_name')
    
    if code:
        drugs = drugs.filter(code__icontains=code)
    if generic_name:
        drugs = drugs.filter(generic_name__icontains=generic_name)
    
    paginator = Paginator(drugs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/yaoqi/drug_list.html', {'page_obj': page_obj})


@login_required
def company_record_submit(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        record_date = request.POST.get('record_date', '').strip()

        if not excel_file:
            messages.error(request, '请选择要上传的Excel文件')
            return redirect('company_record_submit')

        if not record_date:
            messages.error(request, '请选择数据年月')
            return redirect('company_record_submit')

        # 处理年月格式：YYYY-MM → YYYY-MM-DD
        if len(record_date) == 7:
            record_date = record_date + '-01'
        elif len(record_date) == 10:
            pass
        else:
            messages.error(request, f'日期格式不正确：{record_date}')
            return redirect('company_record_submit')

        # 获取当前药企用户关联的企业
        company = None
        if request.user.user_type == 'company':
            try:
                company = request.user.company_profile.company
            except CompanyUser.DoesNotExist:
                company = None

        if not company:
            messages.error(request, '您的账号尚未关联药企，请联系药事所药品科进行药企编码关系维护。')
            return redirect('company_record_submit')

        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            messages.error(request, f'Excel文件读取失败：{str(e)}，请确认文件格式正确')
            return redirect('company_record_submit')

        if len(df) == 0:
            messages.error(request, 'Excel文件中没有数据')
            return redirect('company_record_submit')

        success_count = 0
        skip_count = 0
        error_messages = []
        current_company_name = company.company_name

        for idx, row in df.iterrows():
            row_num = idx + 2
            drug_code = str(row.get('统编代码', '')).strip()
            declared_price = extract_number(row.get('申报价', 0))

            # 校验药企名称：如果Excel中包含企业名称列，则必须与当前药企名称一致
            excel_company_name = row.get('企业名称', row.get('药企名称', ''))
            if excel_company_name and str(excel_company_name).strip():
                excel_name = str(excel_company_name).strip()
                if excel_name != current_company_name:
                    error_messages.append(
                        f'第{row_num}行：企业名称"{excel_name}"与当前登录药企"{current_company_name}"不一致，'
                        f'只能提交本药企的申报价数据'
                    )
                    skip_count += 1
                    continue

            if not drug_code:
                error_messages.append(f'第{row_num}行：统编代码不能为空')
                skip_count += 1
                continue

            drug = Drug.objects.filter(code=drug_code).first()
            if not drug:
                error_messages.append(f'第{row_num}行：统编代码"{drug_code}"对应的药品不存在')
                skip_count += 1
                continue

            if declared_price <= 0:
                error_messages.append(f'第{row_num}行：申报价必须大于0，当前值为{declared_price}')
                skip_count += 1
                continue

            # 创建药企申报价记录
            record, created = CompanyRecord.objects.update_or_create(
                company=company,
                drug=drug,
                record_date=record_date,
                defaults={
                    'declared_price': declared_price,
                    'status': DrugRecordStatus.SUBMITTED,
                }
            )

            # 联动：自动创建药品挂网价审核记录（供药事所药品科审核）
            if drug.network_price > 0 and declared_price != drug.network_price:
                DrugPriceReview.objects.get_or_create(
                    drug=drug,
                    original_price=drug.network_price,
                    proposed_price=declared_price,
                    status=DrugRecordStatus.PENDING,
                    defaults={'comment': f'药企"{current_company_name}"申报，待药品科审核'},
                )

            success_count += 1

        # 汇总反馈
        if success_count > 0:
            messages.success(request, f'成功提交 {success_count} 条申报价记录，已自动提交至药品科审核')
        if skip_count > 0:
            messages.warning(request, f'跳过 {skip_count} 条记录（详见下方提示）')
        if error_messages:
            for msg in error_messages[:20]:
                messages.warning(request, msg)
            if len(error_messages) > 20:
                messages.warning(request, f'...还有 {len(error_messages) - 20} 条错误未显示')
        if success_count == 0 and skip_count == 0:
            messages.error(request, '没有提交任何数据，请检查Excel文件内容')

        return redirect('company_record_list')

    return render(request, 'core/yaoqi/record_submit.html')


@login_required
def company_record_list(request):
    company = None
    if request.user.user_type == 'company':
        try:
            company = request.user.company_profile.company
        except:
            company = None
    
    if company:
        records = CompanyRecord.objects.filter(company=company).select_related('drug')
    else:
        records = CompanyRecord.objects.none()
    
    drug_code = request.GET.get('drug_code')
    status = request.GET.get('status')
    
    if drug_code:
        records = records.filter(drug__code__icontains=drug_code)
    if status:
        records = records.filter(status=status)
    
    paginator = Paginator(records, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/yaoqi/record_list.html', {'page_obj': page_obj})


# ==================== 区县医保局 ====================

@login_required
def district_pharmacy_record_list(request):
    if request.user.user_type == 'district':
        district_name = request.user.organization
    else:
        district_name = request.GET.get('district', '')
    
    records = PharmacyRecord.objects.all().select_related('drug', 'pharmacy').order_by('id')
    
    if district_name:
        records = records.filter(pharmacy__district__icontains=district_name)
    
    drug_code = request.GET.get('drug_code')
    generic_name = request.GET.get('generic_name')
    pharmacy_name = request.GET.get('pharmacy_name')
    pharmacy_code = request.GET.get('pharmacy_code')
    status = request.GET.get('status')
    
    if drug_code:
        records = records.filter(drug__code__icontains=drug_code)
    if generic_name:
        records = records.filter(drug__generic_name__icontains=generic_name)
    if pharmacy_name:
        records = records.filter(pharmacy__pharmacy_name__icontains=pharmacy_name)
    if pharmacy_code:
        records = records.filter(pharmacy__pharmacy_code__icontains=pharmacy_code)
    if status:
        records = records.filter(status=status)
    else:
        records = records.filter(status=DrugRecordStatus.SUBMITTED)
    
    paginator = Paginator(records, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/quxian/record_review.html', {'page_obj': page_obj})


def export_district_pharmacy_records_excel(request):
    if request.user.user_type == 'district':
        district_name = request.user.organization
    else:
        district_name = request.GET.get('district', '')
    
    records = PharmacyRecord.objects.all().select_related('drug', 'pharmacy')
    
    if district_name:
        records = records.filter(pharmacy__district__icontains=district_name)
    
    drug_code = request.GET.get('drug_code')
    generic_name = request.GET.get('generic_name')
    pharmacy_name = request.GET.get('pharmacy_name')
    status = request.GET.get('status')
    
    if drug_code:
        records = records.filter(drug__code__icontains=drug_code)
    if generic_name:
        records = records.filter(drug__generic_name__icontains=generic_name)
    if pharmacy_name:
        records = records.filter(pharmacy__pharmacy_name__icontains=pharmacy_name)
    if status:
        records = records.filter(status=status)
    else:
        records = records.filter(status=DrugRecordStatus.SUBMITTED)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "药店备案价审核"
    
    headers = ['药店名称', '药店代码', '统编代码', '通用名', '上市许可持有人', '备案价', '数据年月', '状态']
    ws.append(headers)
    
    for record in records:
        ws.append([
            record.pharmacy.pharmacy_name,
            record.pharmacy.pharmacy_code,
            record.drug.code,
            record.drug.generic_name,
            record.drug.standard_holder,
            float(record.record_price),
            str(record.record_date),
            record.get_status_display()
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=药店备案价审核_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
def district_pharmacy_record_review(request, pk):
    record = get_object_or_404(PharmacyRecord, pk=pk)
    action = request.POST.get('action')
    comment = request.POST.get('comment', '')
    
    if action == 'approve':
        record.status = DrugRecordStatus.APPROVED
    elif action == 'reject':
        record.status = DrugRecordStatus.REJECTED
    
    record.save()
    
    PharmacyRecordReview.objects.create(
        pharmacy_record=record,
        reviewer=request.user,
        status=record.status,
        comment=comment,
    )
    
    messages.success(request, '审核完成')
    return redirect('district_pharmacy_record_list')


@login_required
def calculate_ratio(request):
    if request.method == 'POST':
        match_catalog = request.POST.get('match_catalog') == 'on'
        match_generic = request.POST.get('match_generic') == 'on'

        if not match_catalog and not match_generic:
            messages.error(request, '请至少选择一种匹配方式')
            return redirect('calculate_ratio')

        drugs = Drug.objects.all()
        drug_list = []
        for drug in drugs:
            drug_list.append({
                'id': drug.id,
                'code': drug.code,
                'drug_category': drug.drug_category,
                'generic_name': drug.generic_name,
                'catalog_dosage_form': drug.catalog_dosage_form,
                'description_dosage_form': drug.description_dosage_form,
                'content': drug.content,
                'volume': drug.volume,
                'quantity': drug.quantity,
                'usage_days': drug.usage_days,
                'standard_holder': drug.standard_holder,
                'catalog_name': drug.catalog_name,
                'special_note': drug.special_note,
                'spec_package': drug.spec_package,
                'network_price': drug.network_price,
            })

        drug_list = process_all_drugs(
            drug_list,
            match_catalog=match_catalog,
            match_generic=match_generic,
        )

        for drug_data in drug_list:
            Drug.objects.filter(id=drug_data['id']).update(
                is_standard=drug_data.get('is_standard', False),
                standard_mark=drug_data.get('standard_mark', ''),
                standard_price=drug_data.get('standard_price', 0),
                price_diff=drug_data.get('price_diff', 0),
                unit_value=drug_data.get('unit_value', 0),
            )

        mode_desc = []
        if match_catalog:
            mode_desc.append('同医保目录')
        if match_generic:
            mode_desc.append('同通用名')
        messages.success(request, f'差比价计算完成（匹配方式：{"+".join(mode_desc)}）')
        return redirect('supervisor_drug_list')

    return render(request, 'core/yaoshisuo/calculate_ratio.html')
